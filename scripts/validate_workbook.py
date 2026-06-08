# -*- coding: utf-8 -*-
"""Validate a v1.0 comparable-pipeline workbook.

Usage:
  python scripts/validate_workbook.py workbook.xlsx
"""

import sys
from pathlib import Path

from openpyxl import load_workbook


ERROR_TOKENS = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"]

COMPARABLE_REQUIRED = [
    "管线/药物", "所属公司/机构", "管线可比性归类", "管线可比性归类说明",
    "管线可比性量化评分（10分）", "命中可比维度", "最高临床阶段/商业状态",
    "所属公司/机构是否有在推进", "是否有在推进的原因", "与目标管线的核心差异", "数据源链接",
]

RAW_REQUIRED = [
    "管线/药物", "所属公司/机构", "管线可比性归类", "管线可比性量化评分（10分）",
    "临床实验的可比性", "临床研究编号/研究名称", "临床期/数据类型", "主要终点", "次要终点",
    "基线", "有效性（分队列/整体，最新披露）", "安全性（分队列/整体，最新披露）",
    "数据源质量和来源小结", "数据源链接",
]

QUANT_REQUIRED = ["模块", "行标题"]


def text(v):
    return "" if v is None else str(v)


def header_map(ws):
    return {text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def missing_columns(ws, required):
    actual = set(header_map(ws))
    return [h for h in required if h not in actual]


def scan_formula_errors(wb):
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(token in cell.value for token in ERROR_TOKENS):
                    hits.append((ws.title, cell.coordinate, cell.value))
    return hits


def nonempty_rows(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            rows.append(r)
    return rows


def scan_required_values(ws, cols):
    hmap = header_map(ws)
    hits = []
    for r in nonempty_rows(ws):
        for col_name in cols:
            c = hmap.get(col_name)
            if c and ws.cell(r, c).value in (None, ""):
                hits.append((ws.title, r, col_name))
    return hits


def validate_scores(ws):
    hmap = header_map(ws)
    col = hmap.get("管线可比性量化评分（10分）")
    if not col:
        return []
    bad = []
    for r in nonempty_rows(ws):
        v = ws.cell(r, col).value
        if v in (None, ""):
            continue
        try:
            score = float(v)
        except (TypeError, ValueError):
            bad.append((ws.title, r, v))
            continue
        if not 0 <= score <= 10:
            bad.append((ws.title, r, v))
    return bad


def validate_hit_dimensions(ws):
    hmap = header_map(ws)
    col = hmap.get("命中可比维度")
    if not col:
        return []
    bad = []
    tokens = ("相同", "不同", "部分相同", "不适用")
    for r in nonempty_rows(ws):
        v = text(ws.cell(r, col).value)
        if v and not any(token in v for token in tokens):
            bad.append((ws.title, r, v))
    return bad


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python validate_workbook.py workbook.xlsx")
    path = Path(sys.argv[1])
    wb = load_workbook(path, data_only=False, rich_text=True)

    print(f"Workbook: {path}")
    print(f"Sheets: {len(wb.sheetnames)}")
    for ws in wb.worksheets:
        print(f"- {ws.title}: {ws.max_row} rows x {ws.max_column} cols")

    errors = scan_formula_errors(wb)
    missing = []
    empty_values = []
    bad_scores = []
    bad_dimensions = []

    for ws in wb.worksheets:
        if ws.title.endswith("的可比管线"):
            missing_cols = missing_columns(ws, COMPARABLE_REQUIRED)
            if missing_cols:
                missing.append((ws.title, missing_cols))
            empty_values += scan_required_values(ws, ["管线/药物", "管线可比性归类", "命中可比维度", "数据源链接"])
            bad_scores += validate_scores(ws)
            bad_dimensions += validate_hit_dimensions(ws)
        elif ws.title.endswith("_量化横比") or ws.title.endswith("_量化横比_精简版"):
            missing_cols = missing_columns(ws, QUANT_REQUIRED)
            if missing_cols:
                missing.append((ws.title, missing_cols))
        elif ws.title != "指标口径" and not ws.title.endswith("_披露总览"):
            missing_cols = missing_columns(ws, RAW_REQUIRED)
            if missing_cols:
                missing.append((ws.title, missing_cols))
            empty_values += scan_required_values(ws, ["管线/药物", "临床研究编号/研究名称", "基线", "有效性（分队列/整体，最新披露）", "安全性（分队列/整体，最新披露）", "数据源链接"])
            bad_scores += validate_scores(ws)
            bad_dimensions += validate_hit_dimensions(ws)

    print(f"Formula-error-like cells: {len(errors)}")
    print(f"Sheets with missing required columns: {len(missing)}")
    print(f"Empty required values: {len(empty_values)}")
    print(f"Invalid scores: {len(bad_scores)}")
    print(f"Unexpected hit-dimension text: {len(bad_dimensions)}")

    for title, cols in missing[:20]:
        print("MISSING_COLUMNS", title, cols)
    for item in empty_values[:20]:
        print("EMPTY_VALUE", item)
    for item in bad_scores[:20]:
        print("BAD_SCORE", item)
    for item in bad_dimensions[:20]:
        print("BAD_DIMENSION", item)
    for item in errors[:20]:
        print("FORMULA_ERROR", item)

    if errors or missing or bad_scores:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
