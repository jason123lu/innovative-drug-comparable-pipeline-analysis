# -*- coding: utf-8 -*-
"""Create a v1.0 comparable-pipeline workbook skeleton from a JSON spec.

Usage:
  python scripts/create_workbook_from_spec.py spec.json output.xlsx
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COMPARABLE_HEADERS = [
    "序号", "管线/药物", "所属公司/机构", "管线可比性归类", "管线可比性归类说明",
    "管线可比性量化评分（10分）", "命中可比维度", "靶点/载荷", "机制", "技术形式/给药",
    "适应症/目标人群", "最高临床阶段/商业状态", "所属公司/机构是否有在推进", "是否有在推进的原因",
    "首次进入临床时间", "最近公开进展时间", "与目标管线的核心差异", "数据源质量", "数据源链接", "备注/待核查点",
]

RAW_HEADERS = [
    "序号", "管线/药物", "所属公司/机构", "管线可比性归类", "管线可比性量化评分（10分）",
    "命中可比维度", "靶点/载荷", "机制", "技术形式/给药", "适应症/目标人群",
    "最高临床阶段/商业状态", "所属公司/机构是否有在推进", "是否有在推进的原因", "与目标管线的核心差异",
    "研究/数据序号", "临床实验的可比性", "临床研究编号/研究名称", "临床期/数据类型",
    "进入该期临床的时间点", "适应症/模型/入组", "主要终点", "次要终点", "基线",
    "有效性（分队列/整体，最新披露）", "安全性（分队列/整体，最新披露）", "上市/标签安全性或黑框",
    "数据源质量和来源小结", "数据源链接", "备注/待核查点",
]

OVERVIEW_HEADERS = [
    "序号", "管线/药物", "管线可比性归类", "最高临床阶段/商业状态",
    "临床前", "临床1期", "临床2期", "临床3期", "上市/标签", "来源/备注",
]

INDICATOR_HEADERS = ["管线/适应症", "最可比有效性指标", "最可比安全性指标", "阶段拆分与比较原则"]

QUANT_HEADERS = ["模块", "行标题"]


FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FILL_SUBHEAD = PatternFill("solid", fgColor="D9EAF7")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, freeze="A2"):
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    for c in range(1, ws.max_column + 1):
        header = str(ws.cell(1, c).value or "")
        width = 16
        if header in {"管线/药物", "所属公司/机构", "管线可比性归类", "靶点/载荷", "机制", "技术形式/给药"}:
            width = 24
        if header in {"管线可比性归类说明", "命中可比维度", "适应症/目标人群", "是否有在推进的原因", "与目标管线的核心差异"}:
            width = 36
        if header in {"基线", "有效性（分队列/整体，最新披露）", "安全性（分队列/整体，最新披露）", "数据源链接"}:
            width = 48
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


def add_sheet(wb, title, headers, freeze="A2"):
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    style_sheet(ws, freeze)
    return ws


def add_comparable_guidance(ws):
    guidance = [
        ["S0", "目标管线本身，作为定位参考。", 10, "相同靶向 / 相同载荷 / 相同技术 / 相同适应症 / 相同终点"],
        ["A", "核心可比：适应症、机制/技术、关键终点高度接近。", 8.5, "相同适应症 / 相同终点 / 部分相同机制"],
        ["B", "终点标尺：机制不同，但可作为疗效或审评参照。", 7, "相同适应症 / 相同终点 / 不同机制"],
    ]
    for row in guidance:
        ws.append([""] + row + [""] * (ws.max_column - 4))


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python create_workbook_from_spec.py spec.json output.xlsx")
    spec_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    spec = json.loads(spec_path.read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)

    if spec.get("make_indicator"):
        add_sheet(wb, "指标口径", INDICATOR_HEADERS)

    for target in spec.get("targets", []):
        target_name = target.get("target_name") or target["sheet_key"].split("_")[0]
        sheet_key = target["sheet_key"]
        comparable_name = target.get("comparable_sheet") or f"{target_name}的可比管线"

        comparable = add_sheet(wb, comparable_name, COMPARABLE_HEADERS, "A2")
        if target.get("include_guidance_rows"):
            add_comparable_guidance(comparable)

        if target.get("make_overview"):
            add_sheet(wb, f"{sheet_key}_披露总览", OVERVIEW_HEADERS, "A2")

        add_sheet(wb, sheet_key, RAW_HEADERS, "A2")

        if target.get("make_quant", True):
            add_sheet(wb, f"{sheet_key}_量化横比", QUANT_HEADERS, "C2")

        if target.get("make_simplified"):
            add_sheet(wb, f"{sheet_key}_量化横比_精简版", QUANT_HEADERS, "C2")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
