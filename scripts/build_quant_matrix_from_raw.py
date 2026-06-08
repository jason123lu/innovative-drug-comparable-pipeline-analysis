# -*- coding: utf-8 -*-
"""Build a quantitative comparison matrix from a research-level raw data sheet.

Usage:
  python scripts/build_quant_matrix_from_raw.py workbook.xlsx raw_sheet [output_sheet]
"""

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASELINE_ROWS = [
    ("样本量/分组", ["样本量", "分组", "n=", "n ", "例", "预计", "入组"]),
    ("年龄/性别", ["年龄", "女性", "男性", "sex", "gender"]),
    ("疾病/风险基线", ["疾病", "风险", "既往", "ASCVD", "ACS", "CAD", "入组", "模型"]),
    ("关键生物标志物基线", ["LDL", "HDL", "TG", "Lp(a)", "ApoB", "hsCRP", "CRP", "PCSK9", "炎症"]),
    ("PAV基线", ["PAV", "percent atheroma"]),
    ("TAV/N-TAV/斑块体积基线", ["TAV", "N-TAV", "normalized", "斑块体积", "plaque volume", "体积"]),
    ("NCPV/LAPV/LAP/NCP基线", ["NCPV", "LAPV", "LAP", "NCP", "低衰减", "非钙化"]),
    ("CIMT/IMT基线", ["CIMT", "IMT", "内膜", "颈动脉"]),
    ("其他影像/病理基线", ["FCT", "LCBI", "脂质弧", "纤维", "巨噬", "坏死", "FAI", "OCT", "IVUS", "CCTA"]),
]

EFFICACY_ROWS = [
    ("PAV变化（绝对值）", ["PAV"], ["绝对", "百分点", "降至", "变化", "回归", "vs"]),
    ("PAV变化（比例）", ["PAV"], ["相对", "比例", "%", "估算", "横比", "降幅"]),
    ("TAV/N-TAV变化（绝对值）", ["TAV", "N-TAV", "Normalized TAV", "斑块体积"], ["绝对", "mm3", "降至", "下降", "变化", "回归", "vs"]),
    ("TAV/N-TAV变化（比例）", ["TAV", "N-TAV", "Normalized TAV", "斑块体积"], ["相对", "比例", "%", "估算", "横比", "降幅"]),
    ("NCPV/LAPV/LAP/NCP变化（绝对值）", ["NCPV", "LAPV", "LAP", "NCP", "低衰减", "非钙化"], []),
    ("NCPV/LAPV/LAP/NCP变化（比例）", ["NCPV", "LAPV", "LAP", "NCP", "低衰减", "非钙化"], ["相对", "比例", "%", "增加", "下降", "变化", "+", "-"]),
    ("CIMT/IMT变化", ["CIMT", "IMT", "内膜", "颈动脉"], []),
    ("纤维帽/FCT/脂质弧/LCBI变化", ["FCT", "纤维帽", "脂质弧", "LCBI", "maxLCBI", "纤维脂质", "MLD"], []),
    ("LDL-C/PCSK9/TG/炎症变化", ["LDL", "HDL", "TG", "Lp(a)", "ApoB", "PCSK9", "hsCRP", "CRP", "炎症", "IL-6"], []),
    ("事件终点/MACE", ["MACE", "事件", "死亡", "MI", "卒中", "血运重建", "survival"], []),
    ("动物斑块面积/体积/病理", ["动物", "小鼠", "兔", "斑块面积", "斑块体积", "病理", "泡沫", "巨噬", "坏死", "saline"], []),
    ("疗效整体结论", [], []),
]

SAFETY_ROWS = [
    ("AE/TEAE/常见不良事件", ["AE", "TEAE", "不良", "常见"]),
    ("TRAE", ["TRAE", "治疗相关"]),
    ("SAE/死亡/停药", ["SAE", "死亡", "停药", "serious"]),
    ("重点安全信号", ["肝", "肌", "CK", "血糖", "注射", "超敏", "输注", "房颤", "出血", "毒性", "感染"]),
    ("上市/标签安全性或黑框", []),
]

BASIC_ROWS = [
    ("管线/药物", "管线/药物"),
    ("所属公司/机构", "所属公司/机构"),
    ("管线可比性归类", "管线可比性归类"),
    ("管线可比性量化评分", "管线可比性量化评分（10分）"),
    ("命中可比维度", "命中可比维度"),
    ("最高临床阶段/商业状态", "最高临床阶段/商业状态"),
    ("所属公司/机构是否有在推进", "所属公司/机构是否有在推进"),
    ("临床实验的可比性", "临床实验的可比性"),
    ("临床研究编号/研究名称", "临床研究编号/研究名称"),
    ("临床期/数据类型", "临床期/数据类型"),
    ("进入该期临床时间", "进入该期临床的时间点"),
    ("适应症/模型/入组", "适应症/模型/入组"),
    ("主要终点", "主要终点"),
    ("次要终点", "次要终点"),
]

SOURCE_ROWS = [
    ("数据源质量和来源小结", "数据源质量和来源小结"),
    ("数据源链接", "数据源链接"),
    ("备注/待核查点", "备注/待核查点"),
]


def text(v):
    return "" if v is None else str(v)


def split_parts(s):
    s = text(s)
    for sep in ["。", "\n"]:
        s = s.replace(sep, "；")
    return [p.strip() for p in s.split("；") if p.strip()]


def has_any(s, words):
    low = s.lower()
    return any(w.lower() in low for w in words)


def pick(parts, include, qualifiers=None, fallback="未披露/不适用"):
    if not include and not qualifiers:
        return "；".join(parts) if parts else fallback
    hits = []
    for part in parts:
        if include and not has_any(part, include):
            continue
        if qualifiers and not has_any(part, qualifiers):
            continue
        hits.append(part)
    return "；".join(dict.fromkeys(hits)) if hits else fallback


def header_map(ws):
    return {text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def row_dict(ws, row, hmap):
    return {name: ws.cell(row, col).value for name, col in hmap.items()}


def style(ws):
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "基础信息": "D9EAF7",
        "疗效": "E2F0D9",
        "安全性": "FCE4D6",
        "备注和来源": "FFF2CC",
    }
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        fill = PatternFill("solid", fgColor=fills.get(text(row[0].value), "FFFFFF"))
        for i, cell in enumerate(row, 1):
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if i > 2 else "left", vertical="top", wrap_text=True)
            if i <= 2:
                cell.fill = fill
                cell.font = Font(bold=True)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22 if c <= 2 else 26


def value_for(raw, label):
    baseline = text(raw.get("基线"))
    efficacy = text(raw.get("有效性（分队列/整体，最新披露）"))
    safety = text(raw.get("安全性（分队列/整体，最新披露）"))

    for output_label, source_col in BASIC_ROWS:
        if label == output_label:
            return raw.get(source_col) or "未披露/不适用"

    for output_label, keywords in BASELINE_ROWS:
        if label == output_label:
            return pick(split_parts(baseline), keywords)

    for output_label, include, qualifiers in EFFICACY_ROWS:
        if label == output_label:
            return pick(split_parts(efficacy), include, qualifiers)

    for output_label, keywords in SAFETY_ROWS:
        if label == output_label:
            if output_label == "上市/标签安全性或黑框":
                return raw.get("上市/标签安全性或黑框") or "未披露/不适用"
            return pick(split_parts(safety), keywords)

    for output_label, source_col in SOURCE_ROWS:
        if label == output_label:
            return raw.get(source_col) or "未披露/不适用"

    return "未披露/不适用"


def main():
    if len(sys.argv) not in (3, 4):
        raise SystemExit("Usage: python build_quant_matrix_from_raw.py workbook.xlsx raw_sheet [output_sheet]")
    path = Path(sys.argv[1])
    raw_sheet = sys.argv[2]
    output_sheet = sys.argv[3] if len(sys.argv) == 4 else f"{raw_sheet}_量化横比"

    wb = load_workbook(path)
    raw_ws = wb[raw_sheet]
    hmap = header_map(raw_ws)

    if output_sheet in wb.sheetnames:
        idx = wb.sheetnames.index(output_sheet)
        wb.remove(wb[output_sheet])
    else:
        idx = wb.sheetnames.index(raw_sheet) + 1
    out = wb.create_sheet(output_sheet[:31], idx)

    raw_rows = [row_dict(raw_ws, r, hmap) for r in range(2, raw_ws.max_row + 1) if raw_ws.cell(r, 1).value not in (None, "")]
    headers = ["模块", "行标题"]
    for raw in raw_rows:
        seq = raw.get("序号") or ""
        data_seq = raw.get("研究/数据序号") or ""
        drug = raw.get("管线/药物") or ""
        study = raw.get("临床研究编号/研究名称") or ""
        headers.append(f"{seq}-{data_seq} {drug}｜{study}")
    out.append(headers)

    row_specs = []
    row_specs += [("基础信息", label) for label, _ in BASIC_ROWS]
    row_specs += [("基础信息", label) for label, _ in BASELINE_ROWS]
    row_specs += [("疗效", label) for label, _, _ in EFFICACY_ROWS]
    row_specs += [("安全性", label) for label, _ in SAFETY_ROWS]
    row_specs += [("备注和来源", label) for label, _ in SOURCE_ROWS]

    for module, label in row_specs:
        out.append([module, label] + [value_for(raw, label) for raw in raw_rows])

    style(out)
    wb.save(path)
    print(path)
    print(f"{output_sheet}: {out.max_row} rows x {out.max_column} cols")


if __name__ == "__main__":
    main()
