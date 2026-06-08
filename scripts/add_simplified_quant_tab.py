# -*- coding: utf-8 -*-
"""Build an optional report-ready simplified tab from a full quant matrix.

Usage:
  python scripts/add_simplified_quant_tab.py workbook.xlsx source_sheet output_sheet
"""

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


KEEP_ROW_TITLES = [
    "管线/药物", "所属公司/机构", "管线可比性归类", "管线可比性量化评分",
    "命中可比维度", "最高临床阶段/商业状态", "所属公司/机构是否有在推进",
    "临床实验的可比性", "临床研究编号/研究名称", "临床期/数据类型",
    "适应症/模型/入组", "主要终点", "次要终点",
    "样本量/分组", "关键生物标志物基线", "影像/病理基线",
    "PAV变化（绝对值）", "PAV变化（比例）",
    "TAV/N-TAV变化（绝对值）", "TAV/N-TAV变化（比例）",
    "NCPV/LAPV/LAP/NCP变化（绝对值）", "NCPV/LAPV/LAP/NCP变化（比例）",
    "CIMT/IMT变化", "纤维帽/FCT/脂质弧/LCBI变化",
    "LDL-C/PCSK9/TG/炎症变化", "事件终点/MACE", "动物斑块面积/体积/病理",
    "AE/TEAE/常见不良事件", "SAE/死亡/停药", "重点安全信号", "上市/标签安全性或黑框",
    "数据源质量和来源小结", "数据源链接", "备注/待核查点",
]


def style(ws):
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "基础信息": "D9EAF7",
        "疗效": "E2F0D9",
        "安全性": "FCE4D6",
        "备注和来源": "FFF2CC",
    }
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        fill = PatternFill("solid", fgColor=fills.get(str(row[0].value or ""), "FFFFFF"))
        for i, cell in enumerate(row, 1):
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if i > 2 else "left", vertical="top", wrap_text=True)
            if i <= 2:
                cell.fill = fill
                cell.font = Font(bold=True)
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22 if c <= 2 else 20
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: python add_simplified_quant_tab.py workbook.xlsx source_sheet output_sheet")
    path = Path(sys.argv[1])
    source_sheet = sys.argv[2]
    output_sheet = sys.argv[3]

    wb = load_workbook(path)
    src = wb[source_sheet]
    if output_sheet in wb.sheetnames:
        idx = wb.sheetnames.index(output_sheet)
        wb.remove(wb[output_sheet])
    else:
        idx = wb.sheetnames.index(source_sheet) + 1
    out = wb.create_sheet(output_sheet[:31], idx)

    keep_cols = [1, 2]
    for c in range(3, src.max_column + 1):
        has_value = any(src.cell(r, c).value not in (None, "", "未披露/不适用") for r in range(2, src.max_row + 1))
        if has_value:
            keep_cols.append(c)
    out.append([src.cell(1, c).value for c in keep_cols])

    keep_set = set(KEEP_ROW_TITLES)
    for r in range(2, src.max_row + 1):
        if str(src.cell(r, 2).value or "") in keep_set:
            out.append([src.cell(r, c).value for c in keep_cols])

    style(out)
    wb.save(path)
    print(path)


if __name__ == "__main__":
    main()
